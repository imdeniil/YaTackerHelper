"""Маршруты для экспорта данных"""

import logging
from datetime import datetime
from io import BytesIO
from fasthtml.common import *
from starlette.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from web.database import get_session, UserCRUD, PaymentRequestCRUD
from web.config import WebConfig
from bot.database.models import UserRole, PaymentRequestStatus
from .decorators import require_auth

logger = logging.getLogger(__name__)


def setup_export_routes(app, config: WebConfig):
    """Настраивает маршруты для экспорта данных"""

    @app.get("/export/excel")
    @require_auth
    async def export_excel(
        sess,
        request,
        search: str = "",
        date_from: str = "",
        date_to: str = "",
        date_type: str = "created",
        amount_min: str = "",
        amount_max: str = "",
        creator_id: int = None,
    ):
        """Экспорт платежей в Excel с учетом фильтров"""
        user_id = sess.get('user_id')
        role = sess.get('role')

        # Извлекаем статусы из query string
        try:
            statuses = request.query_params.getlist('status') if hasattr(request, 'query_params') else []
        except:
            from urllib.parse import parse_qs
            query_string = str(request.url.query) if request.url.query else ""
            query_params = parse_qs(query_string)
            statuses = query_params.get('status', [])

        amount_min_float = float(amount_min) if amount_min else None
        amount_max_float = float(amount_max) if amount_max else None

        async with get_session() as session:
            user = await UserCRUD.get_user_by_id(session, user_id)
            if not user:
                return RedirectResponse('/login', status_code=303)

            # Получаем данные с учетом роли
            if role == UserRole.WORKER.value:
                # Worker видит только свои запросы
                requests_list = await PaymentRequestCRUD.get_payment_requests_advanced(
                    session=session,
                    user_id=user_id,
                    statuses=statuses if statuses else None,
                    search_query=search if search else None,
                    date_from=date_from if date_from else None,
                    date_to=date_to if date_to else None,
                    date_type=date_type,
                    amount_min=amount_min_float,
                    amount_max=amount_max_float,
                    skip=0,
                    limit=0  # Без лимита
                )
            else:
                # Owner/Manager видят все запросы
                requests_list = await PaymentRequestCRUD.get_payment_requests_advanced(
                    session=session,
                    statuses=statuses if statuses else None,
                    search_query=search if search else None,
                    date_from=date_from if date_from else None,
                    date_to=date_to if date_to else None,
                    date_type=date_type,
                    amount_min=amount_min_float,
                    amount_max=amount_max_float,
                    creator_id=creator_id,
                    skip=0,
                    limit=0  # Без лимита
                )

            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Платежи"

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Базовый URL для ссылок на файлы
            base_url = str(request.base_url).rstrip('/')

            # Заголовки
            headers = ["ID", "Название", "Сумма", "Статус", "Создатель", "Дата создания", "Обработал", "Дата оплаты", "Комментарий", "Счёт", "Платёжка"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Маппинг статусов
            status_names = {
                PaymentRequestStatus.PENDING.value: "Ожидает",
                PaymentRequestStatus.SCHEDULED_TODAY.value: "На сегодня",
                PaymentRequestStatus.SCHEDULED_DATE.value: "Запланировано",
                PaymentRequestStatus.PAID.value: "Оплачено",
                PaymentRequestStatus.CANCELLED.value: "Отменено",
            }

            # Стиль для ссылок
            link_font = Font(color="0563C1", underline="single")

            # Данные
            for row_idx, req in enumerate(requests_list, 2):
                ws.cell(row=row_idx, column=1, value=req.id).border = thin_border
                ws.cell(row=row_idx, column=2, value=req.title).border = thin_border
                ws.cell(row=row_idx, column=3, value=req.amount.replace(" ", "")).border = thin_border
                ws.cell(row=row_idx, column=4, value=status_names.get(req.status, req.status)).border = thin_border
                ws.cell(row=row_idx, column=5, value=req.created_by.display_name if req.created_by else "").border = thin_border
                ws.cell(row=row_idx, column=6, value=req.created_at.strftime("%d.%m.%Y %H:%M") if req.created_at else "").border = thin_border
                ws.cell(row=row_idx, column=7, value=req.paid_by.display_name if req.paid_by else "").border = thin_border
                ws.cell(row=row_idx, column=8, value=req.paid_at.strftime("%d.%m.%Y %H:%M") if req.paid_at else "").border = thin_border
                ws.cell(row=row_idx, column=9, value=req.comment or "").border = thin_border

                # Ссылка на счёт
                cell_invoice = ws.cell(row=row_idx, column=10)
                cell_invoice.border = thin_border
                if req.invoice_file_id:
                    cell_invoice.value = "Скачать"
                    cell_invoice.hyperlink = f"{base_url}/payment/{req.id}/download/invoice"
                    cell_invoice.font = link_font

                # Ссылка на платёжку
                cell_proof = ws.cell(row=row_idx, column=11)
                cell_proof.border = thin_border
                if req.payment_proof_file_id and req.payment_proof_file_id != "web_payment":
                    cell_proof.value = "Скачать"
                    cell_proof.hyperlink = f"{base_url}/payment/{req.id}/download/proof"
                    cell_proof.font = link_font

            # Автоширина колонок
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(requests_list) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем в буфер
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Формируем имя файла
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"payments_export_{timestamp}.xlsx"

            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"'
                }
            )
